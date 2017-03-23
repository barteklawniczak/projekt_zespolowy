#########################
#23.03.2017
#Wind speed vs delays
#01.01.2015 analyzed
#########################
import openpyxl
import datetime
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np

#check your paths
flights = "D:\Studia\Semestr VI\Projekt_zespolowy\MyFlights.xlsx"
dublin010115 = "D:\Studia\Semestr VI\Projekt_zespolowy\Dublin010115.xlsx"
fl = openpyxl.load_workbook(flights, read_only=True)
dub = openpyxl.load_workbook(dublin010115, read_only=True)

flsheet = fl['Sheet1']
dubSheet = dub['Dublin010115']

flightTime = []
flightDelay = []
dayTime = []
dayWind = []

for row in range(2,flsheet.max_row):
    if (flsheet._get_cell(row,3).value > datetime.datetime(2015,1,1,0,0)):
        break
    elif (flsheet._get_cell(row,3).value == datetime.datetime(2015,1,1,0,0)):
        flightTime.append(flsheet._get_cell(row,4).value)
        if(flsheet._get_cell(row,15).value == 'NULL'):
            flightDelay.append(0)
        else:
            flightDelay.append(flsheet._get_cell(row,15).value)

for row in range(3,50):         #50-dubSheet max row
    dayWind.append(dubSheet._get_cell(row,9).value)
    dayTime.append(dubSheet._get_cell(row,1).value)

dayWind = [str(x) for x in dayWind]
dWind = [float(i) for i in dayWind]
plt.figure(1)
plt.subplot(211)
plt.plot_date(flightTime, flightDelay)
plt.subplot(212)
plt.plot_date(dayTime,dWind)
plt.show()